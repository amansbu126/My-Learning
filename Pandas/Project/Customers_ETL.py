from openpyxl import load_workbook
from sqlalchemy import create_engine, text
import pymysql

def run_etl():
    EXCEL_FILE = "C:/Users/Aman Kumar/Downloads/Ex_Files_ETL_Python_SQL/Ex_Files_ETL_Python_SQL/Exercise Files/Chapter_2/H+ Sport Customers.xlsx"
    DB_USER = "root"
    DB_PASSWORD = "aman"
    DB_HOST = "localhost"
    DB_PORT = 3306
    DB_NAME = "HR"

    # Step 1: Connect to MySQL
    try:
        engine = create_engine(f"mysql+pymysql://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}")
        with engine.connect() as conn:
            result = conn.execute(text("SELECT NOW();"))
            for row in result:
                print("✅ Connected to MySQL. Current Time:", row[0])
    except Exception as e:
        print("❌ Failed to connect to MySQL:", e)
        return

    # Step 2: Read Excel File (openpyxl)
    try:
        wb = load_workbook(EXCEL_FILE)
        sheet = wb.active
        headers = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))  # ✅ get header values directly
        rows = list(sheet.iter_rows(min_row=2, values_only=True))  # ✅ get all data rows directly

        print(f"🔍 Read {len(rows)} rows from Excel.")
    except Exception as e:
        print("❌ Failed to read Excel file:", e)
        return

    # Step 3: Insert into MySQL
    try:
        # Convert headers to backticked column names for SQL safety
        column_list = ", ".join([f"`{col}`" for col in headers])
        placeholders = ", ".join([f":{col}" for col in headers])
        insert_stmt = text(f"INSERT INTO customers ({column_list}) VALUES ({placeholders})")

        with engine.begin() as conn:
            for row in rows:
                row_dict = dict(zip(headers, row))
                conn.execute(insert_stmt, row_dict)

        print("✅ Data loaded into MySQL table 'customers'.")
    except Exception as e:
        print("❌ Failed to load data into MySQL:", e)

# --- Run ---
if __name__ == "__main__":
    run_etl()

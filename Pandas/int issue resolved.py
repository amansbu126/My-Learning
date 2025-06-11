import openpyxl

# Step 1: Read all lines from the text file
with open("C:/Users/Aman Kumar/OneDrive/문서/New folder/product_key.txt", "r") as file:
    lines = file.readlines()

# Step 2: Process each line
data = []
for line in lines:
    # Split by '|' and strip whitespace, ignore empty entries
    row = [cell.strip() for cell in line.strip().split('|') if cell.strip()]
    data.append(row)

# Step 3: Write to Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Data"

for row_index, row_data in enumerate(data, start=1):
    for col_index, cell_value in enumerate(row_data, start=1):
        try:
            ws.cell(row=row_index, column=col_index).value = int(cell_value)
        except ValueError:
            ws.cell(row=row_index, column=col_index).value = cell_value

# Step 4: Save Excel file
wb.save("C:/Users/Aman Kumar/OneDrive/문서/New folder/product_key_ready.xlsx")

print("Excel file 'output.xlsx' created successfully.")
